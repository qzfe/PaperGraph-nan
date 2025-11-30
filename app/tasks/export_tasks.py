"""导出任务"""
from app.tasks.celery_app import celery_app
from app.database import SessionLocal
from config import settings
import csv, os
from datetime import datetime
from loguru import logger
import openpyxl
from openpyxl.styles import Font, Alignment
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from app.models.mysql_models import (
    PaperInfo, AuthorInfo, OrganizationInfo,
    PaperAuthorRelation, AuthorOrganizationRelation, PaperCitationRelation,
    GraphNodeMapping, StatisticsData, ExportLog
)

@celery_app.task(name="export.generate_file", bind=True)
def generate_export_file(self, export_dao, job_id: str):
    db = SessionLocal()
    
    try:
        job = export_dao.get_job(job_id)
        if not job:
            logger.error(f"任务 {job_id} 不存在")
            return
        
        export_dao.update_status(job_id, "running")
        
        params = job.params
        export_type = params.get("export_type")
        export_format = params.get("format", "csv")
        filters = params.get("filters", {})
        
        data = []
        
        if export_type == "papers":
            paper_dao = PaperDAO(db)
            papers = paper_dao.get_papers(filters)
            data = [
                {
                    "论文ID": p.paper_id,
                    "标题": p.title,
                    "年份": p.year,
                    "会议/期刊": p.venue,
                    "DOI": p.doi,
                    "被引次数": p.citation_count,
                    "关键词": p.keywords
                }
                for p in papers
            ]
        
        elif export_type == "authors":
            author_dao = AuthorDAO(db)
            authors = author_dao.get_authors()
            data = [
                {
                    "作者ID": a.author_id,
                    "姓名": a.name,
                    "单位ID": a.org_id,
                    "H指数": a.h_index,
                    "论文数量": a.paper_count,
                    "ORCID": a.orcid,
                    "邮箱": a.email
                }
                for a in authors
            ]
        
        elif export_type == "organizations":
            org_dao = OrganizationDAO(db)
            orgs = org_dao.get_organizations()
            data = [
                {
                    "单位ID": o.org_id,
                    "名称": o.name,
                    "国家": o.country,
                    "简称": o.abbreviation,
                    "评分": float(o.rank_score) if o.rank_score else 0,
                    "论文数量": o.paper_count
                }
                for o in orgs
            ]
        
        elif export_type == "statistics":
            stats_dao = StatisticsDAO(db)
            stats = stats_dao.query_aggregated(filters)
            data = stats
        
        else:
            raise ValueError(f"不支持的导出类型: {export_type}")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{export_type}_{timestamp}.{export_format}"
        file_path = os.path.join(settings.EXPORT_FILE_DIR, filename)
        
        if export_format == "csv":
            _generate_csv(data, file_path)
        elif export_format == "excel":
            _generate_excel(data, file_path)
        else:
            raise ValueError(f"不支持的导出格式: {export_format}")
        
        export_dao.update_status(job_id, "done", file_path)
        logger.info(f"导出任务 {job_id} 完成: {file_path}")
        
        return {"job_id": job_id, "status": "done", "file_path": file_path}
    
    except Exception as e:
        logger.error(f"导出任务 {job_id} 失败: {e}")
        export_dao.update_status(job_id, "failed", error_msg=str(e))
        raise
    
    finally:
        db.close()

def _generate_csv(data, file_path):
    if not data:
        with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write("")
        return
    
    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

def _generate_excel(data, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据导出"
    
    if not data:
        wb.save(file_path)
        return
    
    headers = list(data[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)

@celery_app.task(name="export.cleanup_old_files")
def cleanup_old_files():
    try:
        export_dir = settings.EXPORT_FILE_DIR
        now = datetime.now()
        
        for filename in os.listdir(export_dir):
            file_path = os.path.join(export_dir, filename)
            
            file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
            age_days = (now - file_time).days
            
            if age_days > 7:
                os.remove(file_path)
                logger.info(f"删除旧文件: {filename}")
        
        logger.info("旧文件清理完成")
    
    except Exception as e:
        logger.error(f"清理旧文件失败: {e}")

class PaperDAO:
    def __init__(self, db: Session):
        self.db = db
    
    def get_papers(self, filters: Dict[str, Any] = None, limit: int = 100) -> List[PaperInfo]:
        query = self.db.query(PaperInfo)
        if filters:
            if "year" in filters:
                query = query.filter(PaperInfo.year == filters["year"])
            if "keyword" in filters:
                query = query.filter(PaperInfo.keywords.like(f"%{filters['keyword']}%"))
        return query.limit(limit).all()

class AuthorDAO:
    def __init__(self, db: Session):
        self.db = db
    
    def get_authors(self, limit: int = 100) -> List[AuthorInfo]:
        return self.db.query(AuthorInfo).limit(limit).all()

class OrganizationDAO:
    def __init__(self, db: Session):
        self.db = db
    
    def get_organizations(self, limit: int = 100) -> List[OrganizationInfo]:
        return self.db.query(OrganizationInfo).limit(limit).all()
    
class StatisticsDAO:
    def __init__(self, db: Session):
        self.db = db
    
    def query_aggregated(self, query_params: Dict[str, Any]) -> List[Dict[str, Any]]:
        metric = query_params.get("metric")
        start_year = query_params.get("start_year")
        end_year = query_params.get("end_year")
        limit = query_params.get("limit", 100)
        
        try:
            if metric == "paper_count_by_year":
                query = self.db.query(
                    PaperInfo.year,
                    func.count(PaperInfo.paper_id).label("count")
                ).group_by(PaperInfo.year)
                
                if start_year:
                    query = query.filter(PaperInfo.year >= start_year)
                if end_year:
                    query = query.filter(PaperInfo.year <= end_year)
                
                query = query.order_by(PaperInfo.year).limit(limit)
                results = query.all()
                return [{"label": str(r.year), "value": r.count} for r in results]
            
            elif metric == "top_authors":
                query = self.db.query(
                    AuthorInfo.author_id,
                    AuthorInfo.name,
                    AuthorInfo.paper_count,
                    AuthorInfo.h_index
                ).order_by(AuthorInfo.paper_count.desc()).limit(limit)
                
                results = query.all()
                return [{"label": r.name, "value": r.paper_count, "extra": {"h_index": r.h_index, "author_id": r.author_id}} for r in results]
            
            elif metric == "top_organizations":
                query = self.db.query(
                    OrganizationInfo.org_id,
                    OrganizationInfo.name,
                    OrganizationInfo.paper_count,
                    OrganizationInfo.rank_score
                ).order_by(OrganizationInfo.paper_count.desc()).limit(limit)
                
                results = query.all()
                return [{"label": r.name, "value": r.paper_count, "extra": {"rank_score": float(r.rank_score) if r.rank_score else 0, "org_id": r.org_id}} for r in results]
            
            else:
                query = self.db.query(StatisticsData).filter(StatisticsData.metric == metric).limit(limit)
                results = query.all()
                return [{"label": r.dims_json.get("label", ""), "value": float(r.value) if r.value else 0, "extra": r.dims_json} for r in results]
        
        except Exception as e:
            logger.error(f"查询统计数据失败: {e}")
            raise